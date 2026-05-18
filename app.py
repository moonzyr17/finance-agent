import os
import json
import sqlite3
import requests
import base64
from datetime import datetime, date
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file, g
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# On Vercel/serverless, only /tmp is writable. Use /tmp on read-only filesystems.
IS_SERVERLESS = bool(os.getenv('VERCEL') or os.getenv('AWS_LAMBDA_FUNCTION_NAME'))
if IS_SERVERLESS:
    DB_PATH = '/tmp/finance.db'
    UPLOAD_FOLDER = '/tmp/uploads'
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
else:
    DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'finance.db')
OPENROUTER_API_KEY = os.getenv('OPENROUTER_API_KEY', '')

# Seed data for fresh deployments (so demo always has data)
SEED_INVOICES = [
    {'invoice_no':'INV-2026-001','vendor':'Notion Labs Inc','amount':500000,'currency':'IDR','issue_date':'2026-05-01','due_date':'2026-05-15','category':'Software','status':'paid','description':'Team workspace subscription'},
    {'invoice_no':'INV-2026-002','vendor':'Google Workspace','amount':1200000,'currency':'IDR','issue_date':'2026-05-03','due_date':'2026-05-25','category':'Software','status':'paid','description':'Business Standard plan'},
    {'invoice_no':'INV-2026-003','vendor':'Tokopedia Office','amount':250000,'currency':'IDR','issue_date':'2026-05-05','due_date':'2026-05-20','category':'Office Supplies','status':'paid','description':'Stationery and printer ink'},
    {'invoice_no':'INV-2026-004','vendor':'Grab Business','amount':350000,'currency':'IDR','issue_date':'2026-05-10','due_date':'2026-05-22','category':'Travel','status':'unpaid','description':'Client meeting transport'},
    {'invoice_no':'INV-2026-005','vendor':'PLN Persero','amount':800000,'currency':'IDR','issue_date':'2026-05-12','due_date':'2026-05-23','category':'Utilities','status':'unpaid','description':'May electricity bill'},
    {'invoice_no':'INV-2026-006','vendor':'Facebook Ads','amount':2000000,'currency':'IDR','issue_date':'2026-05-15','due_date':'2026-05-30','category':'Marketing','status':'unpaid','description':'Q2 campaign boost'},
    {'invoice_no':'INV-2026-007','vendor':'Freelance Designer','amount':1500000,'currency':'IDR','issue_date':'2026-05-18','due_date':'2026-06-01','category':'Services','status':'unpaid','description':'Logo redesign project'},
    {'invoice_no':'INV-2026-008','vendor':'Indihome Telkom','amount':450000,'currency':'IDR','issue_date':'2026-04-25','due_date':'2026-05-10','category':'Utilities','status':'overdue','description':'Office internet April'},
    {'invoice_no':'INV-2026-009','vendor':'AWS Cloud','amount':3200000,'currency':'IDR','issue_date':'2026-05-08','due_date':'2026-05-28','category':'Software','status':'unpaid','description':'EC2 + S3 monthly'},
    {'invoice_no':'INV-2026-010','vendor':'Canva Pro','amount':180000,'currency':'IDR','issue_date':'2026-05-02','due_date':'2026-05-17','category':'Software','status':'paid','description':'Pro team license'},
]

# ─── DB ───────────────────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    db.executescript('''
        CREATE TABLE IF NOT EXISTS invoices (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_no  TEXT,
            vendor      TEXT,
            amount      REAL,
            currency    TEXT DEFAULT 'IDR',
            issue_date  TEXT,
            due_date    TEXT,
            status      TEXT DEFAULT 'unpaid',
            category    TEXT,
            description TEXT,
            image_path  TEXT,
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS transactions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT,
            description TEXT,
            amount      REAL,
            type        TEXT,
            reference   TEXT,
            matched_invoice_id INTEGER,
            created_at  TEXT DEFAULT (datetime('now'))
        );
    ''')
    # Seed demo data only if empty
    count = db.execute('SELECT COUNT(*) FROM invoices').fetchone()[0]
    if count == 0:
        for s in SEED_INVOICES:
            db.execute(
                '''INSERT INTO invoices (invoice_no, vendor, amount, currency, issue_date, due_date, status, category, description)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                (s['invoice_no'], s['vendor'], s['amount'], s['currency'],
                 s['issue_date'], s['due_date'], s['status'], s['category'], s['description'])
            )
    db.commit()
    db.close()

# ─── GEMINI FLASH VISION ──────────────────────────────────────────────────────

def parse_invoice_with_gemini(image_bytes: bytes, mime_type: str) -> dict:
    b64 = base64.b64encode(image_bytes).decode()
    prompt = """You are a finance assistant. Extract structured data from this invoice image.
Return ONLY valid JSON with these fields (use null if not found):
{
  "invoice_no": "string",
  "vendor": "string",
  "amount": number,
  "currency": "IDR or USD or etc",
  "issue_date": "YYYY-MM-DD",
  "due_date": "YYYY-MM-DD",
  "category": "one of: Software, Office Supplies, Travel, Marketing, Utilities, Services, Other",
  "description": "brief description of what was purchased"
}
Do not include markdown, only raw JSON."""

    payload = {
        "model": "google/gemini-2.5-flash",
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}}
                ]
            }
        ],
        "max_tokens": 512
    }

    resp = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json"
        },
        json=payload,
        timeout=30
    )
    resp.raise_for_status()
    content = resp.json()['choices'][0]['message']['content'].strip()
    # strip markdown code fences if present
    if content.startswith('```'):
        content = content.split('```')[1]
        if content.startswith('json'):
            content = content[4:]
    return json.loads(content.strip())

# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/invoices')
def invoices_page():
    return render_template('invoices.html')

@app.route('/reconcile')
def reconcile_page():
    return render_template('reconcile.html')

@app.route('/reports')
def reports_page():
    return render_template('reports.html')

# ── API: invoices ──

@app.route('/api/invoices', methods=['GET'])
def get_invoices():
    db = get_db()
    status = request.args.get('status')
    q = 'SELECT * FROM invoices'
    params = []
    if status:
        q += ' WHERE status = ?'
        params.append(status)
    q += ' ORDER BY due_date ASC'
    rows = db.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/invoices', methods=['POST'])
def create_invoice():
    data = request.json
    db = get_db()
    cur = db.execute(
        '''INSERT INTO invoices (invoice_no, vendor, amount, currency, issue_date, due_date, status, category, description)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (data.get('invoice_no'), data.get('vendor'), data.get('amount'),
         data.get('currency', 'IDR'), data.get('issue_date'), data.get('due_date'),
         data.get('status', 'unpaid'), data.get('category'), data.get('description'))
    )
    db.commit()
    return jsonify({'id': cur.lastrowid, 'message': 'Invoice created'}), 201

@app.route('/api/invoices/<int:inv_id>', methods=['PATCH'])
def update_invoice(inv_id):
    data = request.json
    db = get_db()
    fields = []
    vals = []
    for k in ('status', 'vendor', 'amount', 'due_date', 'category', 'description'):
        if k in data:
            fields.append(f'{k} = ?')
            vals.append(data[k])
    if not fields:
        return jsonify({'error': 'No fields to update'}), 400
    vals.append(inv_id)
    db.execute(f'UPDATE invoices SET {", ".join(fields)} WHERE id = ?', vals)
    db.commit()
    return jsonify({'message': 'Updated'})

@app.route('/api/invoices/<int:inv_id>', methods=['DELETE'])
def delete_invoice(inv_id):
    db = get_db()
    db.execute('DELETE FROM invoices WHERE id = ?', (inv_id,))
    db.commit()
    return jsonify({'message': 'Deleted'})

# ── API: scan invoice image ──

@app.route('/api/scan', methods=['POST'])
def scan_invoice():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Empty filename'}), 400

    image_bytes = f.read()
    mime_type = f.content_type or 'image/jpeg'

    if not OPENROUTER_API_KEY:
        return jsonify({'error': 'OPENROUTER_API_KEY not set'}), 500

    try:
        parsed = parse_invoice_with_gemini(image_bytes, mime_type)
    except Exception as e:
        return jsonify({'error': f'Gemini parse failed: {str(e)}'}), 500

    # save image
    ext = f.filename.rsplit('.', 1)[-1] if '.' in f.filename else 'jpg'
    fname = f"invoice_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    with open(fpath, 'wb') as fp:
        fp.write(image_bytes)
    parsed['image_path'] = f'/static/uploads/{fname}'

    return jsonify(parsed)

# ── API: stats ──

@app.route('/api/stats')
def get_stats():
    db = get_db()
    total = db.execute('SELECT COALESCE(SUM(amount),0) FROM invoices').fetchone()[0]
    paid = db.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='paid'").fetchone()[0]
    unpaid = db.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='unpaid'").fetchone()[0]
    overdue = db.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='overdue'").fetchone()[0]
    count = db.execute('SELECT COUNT(*) FROM invoices').fetchone()[0]
    overdue_count = db.execute("SELECT COUNT(*) FROM invoices WHERE status='overdue'").fetchone()[0]

    # auto-mark overdue
    today = date.today().isoformat()
    db.execute("UPDATE invoices SET status='overdue' WHERE status='unpaid' AND due_date < ?", (today,))
    db.commit()

    return jsonify({
        'total': total, 'paid': paid, 'unpaid': unpaid, 'overdue': overdue,
        'count': count, 'overdue_count': overdue_count
    })

# ── API: transactions (for reconciliation) ──

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    db = get_db()
    rows = db.execute('SELECT * FROM transactions ORDER BY date DESC').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/transactions/upload', methods=['POST'])
def upload_transactions():
    """Accept CSV: date,description,amount,type,reference"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    import csv, io
    content = request.files['file'].read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    db = get_db()
    inserted = 0
    for row in reader:
        db.execute(
            'INSERT INTO transactions (date, description, amount, type, reference) VALUES (?,?,?,?,?)',
            (row.get('date'), row.get('description'), float(row.get('amount', 0)),
             row.get('type', 'debit'), row.get('reference', ''))
        )
        inserted += 1
    db.commit()
    return jsonify({'inserted': inserted})

@app.route('/api/reconcile', methods=['POST'])
def reconcile():
    """Match transactions to invoices by amount proximity."""
    db = get_db()
    invoices = db.execute("SELECT * FROM invoices WHERE status != 'paid'").fetchall()
    transactions = db.execute('SELECT * FROM transactions WHERE matched_invoice_id IS NULL').fetchall()

    matched = []
    unmatched_inv = []
    unmatched_tx = []

    used_tx = set()
    for inv in invoices:
        found = None
        for tx in transactions:
            if tx['id'] in used_tx:
                continue
            if abs(tx['amount'] - inv['amount']) / max(inv['amount'], 1) < 0.01:
                found = tx
                break
        if found:
            used_tx.add(found['id'])
            db.execute('UPDATE transactions SET matched_invoice_id=? WHERE id=?', (inv['id'], found['id']))
            matched.append({'invoice': dict(inv), 'transaction': dict(found)})
        else:
            unmatched_inv.append(dict(inv))

    for tx in transactions:
        if tx['id'] not in used_tx:
            unmatched_tx.append(dict(tx))

    db.commit()
    return jsonify({'matched': matched, 'unmatched_invoices': unmatched_inv, 'unmatched_transactions': unmatched_tx})

# ── API: monthly report ──

@app.route('/api/report/monthly')
def monthly_report():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    db = get_db()
    rows = db.execute(
        "SELECT * FROM invoices WHERE strftime('%Y-%m', issue_date) = ? ORDER BY issue_date",
        (month,)
    ).fetchall()
    data = [dict(r) for r in rows]
    total = sum(r['amount'] for r in data)
    by_cat = {}
    for r in data:
        cat = r['category'] or 'Other'
        by_cat[cat] = by_cat.get(cat, 0) + r['amount']
    return jsonify({'month': month, 'invoices': data, 'total': total, 'by_category': by_cat})

# ── API: export Excel ──

@app.route('/api/export/excel')
def export_excel():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    db = get_db()
    rows = db.execute(
        "SELECT * FROM invoices WHERE strftime('%Y-%m', issue_date) = ? ORDER BY issue_date",
        (month,)
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {month}"

    # header style
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin', color='E4E7EB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['No', 'Invoice No', 'Vendor', 'Category', 'Issue Date', 'Due Date', 'Amount', 'Currency', 'Status', 'Description']
    col_widths = [5, 15, 20, 15, 12, 12, 15, 10, 10, 30]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28

    status_colors = {'paid': 'D1FAE5', 'unpaid': 'FEF3C7', 'overdue': 'FEE2E2'}

    total = 0
    for idx, row in enumerate(rows, 2):
        vals = [idx-1, row['invoice_no'], row['vendor'], row['category'],
                row['issue_date'], row['due_date'], row['amount'],
                row['currency'], row['status'], row['description']]
        fill_color = status_colors.get(row['status'], 'FFFFFF')
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=fill_color)
            if col == 7:  # amount
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        total += row['amount'] or 0

    # total row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=6, value='TOTAL').font = Font(bold=True)
    tc = ws.cell(row=total_row, column=7, value=total)
    tc.font = Font(bold=True)
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal='right')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'finance-report-{month}.xlsx')

# ── API: reminders ──

@app.route('/api/reminders')
def get_reminders():
    db = get_db()
    today = date.today().isoformat()
    # invoices due in next 7 days or already overdue
    rows = db.execute(
        """SELECT * FROM invoices
           WHERE status != 'paid'
           AND (due_date <= date('now', '+7 days'))
           ORDER BY due_date ASC""",
    ).fetchall()
    return jsonify([dict(r) for r in rows])

# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5099)

with app.app_context():
    init_db()
